"""Document Analysis System for Jarvis/Ultron.

This module provides comprehensive document parsing and understanding:
- PDF analysis (text, tables, structure)
- Spreadsheet parsing (Excel, CSV)
- Image document OCR (via Claude Vision)
- AI-powered summarization and data extraction
- Document comparison and search

Supports Jarvis (interactive analysis) and Ultron (background processing).
"""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import logging
import mimetypes
import os
import re
import shutil
import tempfile
from dataclasses import dataclass, field, asdict
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from uuid import uuid4

from anthropic import AsyncAnthropic

from app.core.config import settings

logger = logging.getLogger(__name__)


# ============ ENUMS AND DATA CLASSES ============

class DocumentType(str, Enum):
    """Supported document types."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    IMAGE = "image"
    TEXT = "text"
    MARKDOWN = "markdown"
    UNKNOWN = "unknown"


class ExtractionType(str, Enum):
    """Types of data extraction."""
    DATES = "dates"
    AMOUNTS = "amounts"
    NAMES = "names"
    EMAILS = "emails"
    PHONES = "phones"
    ADDRESSES = "addresses"
    TABLES = "tables"
    CUSTOM = "custom"


@dataclass
class DocumentMetadata:
    """Metadata about a processed document."""
    id: str = ""
    filename: str = ""
    document_type: DocumentType = DocumentType.UNKNOWN
    file_size: int = 0
    page_count: int = 0
    created_at: str = ""
    processed_at: str = ""

    # Content info
    has_text: bool = True
    has_tables: bool = False
    has_images: bool = False
    language: str = "en"

    # Processing info
    processing_time_ms: int = 0
    tokens_used: int = 0

    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        data["document_type"] = self.document_type.value
        return data


@dataclass
class ExtractedTable:
    """A table extracted from a document."""
    headers: List[str] = field(default_factory=list)
    rows: List[List[str]] = field(default_factory=list)
    page_number: int = 1
    table_index: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def to_markdown(self) -> str:
        """Convert table to markdown format."""
        if not self.headers and not self.rows:
            return ""

        lines = []

        # Headers
        if self.headers:
            lines.append("| " + " | ".join(self.headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(self.headers)) + " |")

        # Rows
        for row in self.rows:
            lines.append("| " + " | ".join(str(cell) for cell in row) + " |")

        return "\n".join(lines)


@dataclass
class DocumentAnalysis:
    """Result of analyzing a document."""
    success: bool = True
    error: Optional[str] = None

    # Metadata
    metadata: Optional[DocumentMetadata] = None

    # Content
    text: str = ""
    pages: List[str] = field(default_factory=list)  # Text per page
    tables: List[ExtractedTable] = field(default_factory=list)

    # AI-generated content
    summary: str = ""
    key_points: List[str] = field(default_factory=list)
    extracted_data: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        data = {
            "success": self.success,
            "error": self.error,
            "metadata": self.metadata.to_dict() if self.metadata else None,
            "text": self.text,
            "pages": self.pages,
            "tables": [t.to_dict() for t in self.tables],
            "summary": self.summary,
            "key_points": self.key_points,
            "extracted_data": self.extracted_data,
        }
        return data


# ============ DOCUMENT STORE ============

class DocumentStore:
    """Secure temporary file storage for document processing.

    Features:
    - Secure temp directory with auto-cleanup
    - File type validation
    - Size limits
    - Automatic expiration
    """

    # Maximum file size (50MB)
    MAX_FILE_SIZE = 50 * 1024 * 1024

    # Allowed file extensions
    ALLOWED_EXTENSIONS = {
        ".pdf", ".xlsx", ".xls", ".csv",
        ".jpg", ".jpeg", ".png", ".gif", ".webp", ".tiff", ".bmp",
        ".txt", ".md", ".markdown", ".text",
    }

    # MIME type mapping
    MIME_TYPES = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".tiff": "image/tiff",
        ".bmp": "image/bmp",
        ".txt": "text/plain",
        ".md": "text/markdown",
        ".markdown": "text/markdown",
    }

    def __init__(self, base_dir: Optional[str] = None):
        """Initialize document store.

        Args:
            base_dir: Base directory for temp files. Uses system temp if not provided.
        """
        self.base_dir = Path(base_dir) if base_dir else Path(tempfile.gettempdir())
        self.store_dir = self.base_dir / "nexus_documents"
        self._ensure_store_dir()

        # Track stored files
        self._files: Dict[str, Dict[str, Any]] = {}

    def _ensure_store_dir(self):
        """Ensure the store directory exists with proper permissions."""
        self.store_dir.mkdir(parents=True, exist_ok=True)
        # Restrict permissions to owner only
        try:
            os.chmod(self.store_dir, 0o700)
        except OSError:
            pass  # May fail on some systems

    def validate_file(self, filename: str, content: bytes) -> Tuple[bool, Optional[str]]:
        """Validate a file for storage.

        Args:
            filename: Original filename
            content: File content as bytes

        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check file size
        if len(content) > self.MAX_FILE_SIZE:
            return False, f"File too large. Maximum size: {self.MAX_FILE_SIZE / 1024 / 1024:.0f}MB"

        # Check extension
        ext = Path(filename).suffix.lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            return False, f"File type not allowed: {ext}. Allowed: {', '.join(sorted(self.ALLOWED_EXTENSIONS))}"

        # Basic magic number validation for common types
        if not self._validate_magic_bytes(ext, content):
            return False, "File content does not match extension"

        return True, None

    def _validate_magic_bytes(self, ext: str, content: bytes) -> bool:
        """Validate file magic bytes match expected type."""
        if len(content) < 8:
            return True  # Too small to validate

        magic_bytes = {
            ".pdf": [b"%PDF"],
            ".xlsx": [b"PK\x03\x04"],  # ZIP format
            ".xls": [b"\xd0\xcf\x11\xe0"],  # OLE format
            ".png": [b"\x89PNG"],
            ".jpg": [b"\xff\xd8\xff"],
            ".jpeg": [b"\xff\xd8\xff"],
            ".gif": [b"GIF87a", b"GIF89a"],
            ".webp": [b"RIFF"],  # WebP is RIFF format
        }

        if ext in magic_bytes:
            return any(content.startswith(magic) for magic in magic_bytes[ext])

        return True  # No magic bytes to check

    def store_file(
        self,
        filename: str,
        content: bytes,
        user_id: Optional[str] = None,
    ) -> Tuple[str, str]:
        """Store a file and return its ID and path.

        Args:
            filename: Original filename
            content: File content
            user_id: Optional user ID for tracking

        Returns:
            Tuple of (file_id, file_path)

        Raises:
            ValueError: If file validation fails
        """
        # Validate
        is_valid, error = self.validate_file(filename, content)
        if not is_valid:
            raise ValueError(error)

        # Generate unique ID
        file_id = f"{uuid4().hex[:16]}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

        # Sanitize filename
        safe_name = self._sanitize_filename(filename)

        # Create file path
        file_path = self.store_dir / f"{file_id}_{safe_name}"

        # Write file
        with open(file_path, "wb") as f:
            f.write(content)

        # Track file
        self._files[file_id] = {
            "path": str(file_path),
            "original_name": filename,
            "size": len(content),
            "stored_at": datetime.utcnow().isoformat(),
            "user_id": user_id,
        }

        logger.info(f"Stored document: {file_id} ({len(content)} bytes)")

        return file_id, str(file_path)

    def store_from_base64(
        self,
        filename: str,
        base64_content: str,
        user_id: Optional[str] = None,
    ) -> Tuple[str, str]:
        """Store a file from base64 encoded content.

        Args:
            filename: Original filename
            base64_content: Base64 encoded content (may include data URL prefix)
            user_id: Optional user ID

        Returns:
            Tuple of (file_id, file_path)
        """
        # Remove data URL prefix if present
        if base64_content.startswith("data:"):
            base64_content = base64_content.split(",", 1)[1]

        # Decode
        content = base64.b64decode(base64_content)

        return self.store_file(filename, content, user_id)

    def get_file(self, file_id: str) -> Optional[Dict[str, Any]]:
        """Get file info by ID.

        Args:
            file_id: File ID

        Returns:
            File info dict or None if not found
        """
        return self._files.get(file_id)

    def get_file_path(self, file_id: str) -> Optional[str]:
        """Get file path by ID.

        Args:
            file_id: File ID

        Returns:
            File path or None if not found
        """
        info = self._files.get(file_id)
        return info["path"] if info else None

    def read_file(self, file_id: str) -> Optional[bytes]:
        """Read file content by ID.

        Args:
            file_id: File ID

        Returns:
            File content or None if not found
        """
        path = self.get_file_path(file_id)
        if not path or not Path(path).exists():
            return None

        with open(path, "rb") as f:
            return f.read()

    def delete_file(self, file_id: str) -> bool:
        """Delete a stored file.

        Args:
            file_id: File ID

        Returns:
            True if deleted, False if not found
        """
        info = self._files.pop(file_id, None)
        if not info:
            return False

        path = Path(info["path"])
        if path.exists():
            path.unlink()
            logger.info(f"Deleted document: {file_id}")
            return True

        return False

    def cleanup_old_files(self, max_age_hours: int = 24):
        """Remove files older than specified age.

        Args:
            max_age_hours: Maximum age in hours before cleanup
        """
        cutoff = datetime.utcnow().timestamp() - (max_age_hours * 3600)

        to_delete = []
        for file_id, info in self._files.items():
            stored_at = datetime.fromisoformat(info["stored_at"])
            if stored_at.timestamp() < cutoff:
                to_delete.append(file_id)

        for file_id in to_delete:
            self.delete_file(file_id)

        if to_delete:
            logger.info(f"Cleaned up {len(to_delete)} old documents")

    def cleanup_all(self):
        """Remove all stored files."""
        for file_id in list(self._files.keys()):
            self.delete_file(file_id)

        # Also clean store directory
        if self.store_dir.exists():
            for item in self.store_dir.iterdir():
                if item.is_file():
                    item.unlink()

    def _sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for safe storage."""
        # Remove path components
        name = Path(filename).name

        # Replace dangerous characters
        name = re.sub(r'[^\w\s.-]', '_', name)

        # Limit length
        if len(name) > 100:
            ext = Path(name).suffix
            name = name[:100-len(ext)] + ext

        return name


# ============ DOCUMENT ANALYZER ============

class DocumentAnalyzer:
    """Analyzes and extracts information from various document types.

    Supports:
    - PDF (text and structure extraction)
    - Excel/CSV (spreadsheet parsing)
    - Images (OCR via Claude Vision)
    - Plain text and Markdown

    Uses Claude for AI-powered analysis and extraction.
    """

    # Model for document analysis
    ANALYSIS_MODEL = "claude-sonnet-4-20250514"
    VISION_MODEL = "claude-sonnet-4-20250514"
    MAX_TOKENS = 4096

    # Prompts
    SUMMARY_PROMPT = """Analyze this document and provide:

1. **Executive Summary** (2-3 sentences): What is this document about?
2. **Key Points** (bullet list): The most important information
3. **Document Type**: What kind of document is this (invoice, contract, report, etc.)?
4. **Key Dates**: Any important dates mentioned
5. **Key Amounts**: Any monetary values or quantities
6. **People/Organizations**: Names of people or companies mentioned

DOCUMENT CONTENT:
{content}

Provide your analysis in a clear, structured format."""

    EXTRACTION_PROMPT = """Extract the following information from this document:

{extraction_type}

Return the extracted data as a JSON object. If a field cannot be found, use null.
Be precise and extract exactly what's in the document.

DOCUMENT CONTENT:
{content}

EXTRACTED DATA (JSON):"""

    SEARCH_PROMPT = """Search this document for information related to: "{query}"

Return:
1. All relevant passages that mention or relate to the query
2. Page numbers or locations if available
3. Context around each match

DOCUMENT CONTENT:
{content}

SEARCH RESULTS:"""

    COMPARE_PROMPT = """Compare these two document versions and identify:

1. **Added Content**: New text in the second version
2. **Removed Content**: Text that was in the first but not the second
3. **Modified Content**: Text that changed between versions
4. **Unchanged Sections**: Major sections that stayed the same
5. **Summary of Changes**: Brief overview of what changed

DOCUMENT VERSION 1:
{content1}

DOCUMENT VERSION 2:
{content2}

COMPARISON RESULTS:"""

    def __init__(
        self,
        client: Optional[AsyncAnthropic] = None,
        document_store: Optional[DocumentStore] = None,
    ):
        """Initialize document analyzer.

        Args:
            client: Anthropic client. Creates one if not provided.
            document_store: Document store for file handling.
        """
        self.client = client or AsyncAnthropic(api_key=settings.anthropic_api_key)
        self.store = document_store or DocumentStore()

    def detect_document_type(self, filename: str, content: Optional[bytes] = None) -> DocumentType:
        """Detect document type from filename and optionally content.

        Args:
            filename: Filename with extension
            content: Optional file content for magic byte detection

        Returns:
            Detected DocumentType
        """
        ext = Path(filename).suffix.lower()

        type_mapping = {
            ".pdf": DocumentType.PDF,
            ".xlsx": DocumentType.EXCEL,
            ".xls": DocumentType.EXCEL,
            ".csv": DocumentType.CSV,
            ".jpg": DocumentType.IMAGE,
            ".jpeg": DocumentType.IMAGE,
            ".png": DocumentType.IMAGE,
            ".gif": DocumentType.IMAGE,
            ".webp": DocumentType.IMAGE,
            ".tiff": DocumentType.IMAGE,
            ".bmp": DocumentType.IMAGE,
            ".txt": DocumentType.TEXT,
            ".md": DocumentType.MARKDOWN,
            ".markdown": DocumentType.MARKDOWN,
        }

        return type_mapping.get(ext, DocumentType.UNKNOWN)

    async def analyze_pdf(
        self,
        file_path: str,
        extract_tables: bool = True,
        max_pages: Optional[int] = None,
    ) -> DocumentAnalysis:
        """Extract text, tables, and structure from a PDF.

        Args:
            file_path: Path to PDF file
            extract_tables: Whether to attempt table extraction
            max_pages: Maximum pages to process (None = all)

        Returns:
            DocumentAnalysis with extracted content
        """
        import time
        start_time = time.time()

        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return DocumentAnalysis(
                success=False,
                error=f"File not found: {file_path}"
            )

        try:
            # Try pdfplumber first (better for tables)
            try:
                import pdfplumber
                return await self._analyze_pdf_with_pdfplumber(
                    path, extract_tables, max_pages, start_time
                )
            except ImportError:
                pass

            # Fallback to PyPDF2
            try:
                import PyPDF2
                return await self._analyze_pdf_with_pypdf2(
                    path, max_pages, start_time
                )
            except ImportError:
                pass

            return DocumentAnalysis(
                success=False,
                error="No PDF library available. Install pdfplumber or PyPDF2."
            )

        except Exception as e:
            logger.error(f"PDF analysis failed: {e}")
            return DocumentAnalysis(
                success=False,
                error=f"PDF analysis failed: {str(e)}"
            )

    async def _analyze_pdf_with_pdfplumber(
        self,
        path: Path,
        extract_tables: bool,
        max_pages: Optional[int],
        start_time: float,
    ) -> DocumentAnalysis:
        """Analyze PDF using pdfplumber."""
        import pdfplumber
        import time

        pages_text = []
        tables = []

        with pdfplumber.open(path) as pdf:
            page_count = len(pdf.pages)
            pages_to_process = pdf.pages[:max_pages] if max_pages else pdf.pages

            for page_idx, page in enumerate(pages_to_process):
                # Extract text
                text = page.extract_text() or ""
                pages_text.append(text)

                # Extract tables
                if extract_tables:
                    page_tables = page.extract_tables() or []
                    for table_idx, table_data in enumerate(page_tables):
                        if table_data and len(table_data) > 0:
                            headers = [str(cell) if cell else "" for cell in table_data[0]]
                            rows = [
                                [str(cell) if cell else "" for cell in row]
                                for row in table_data[1:]
                            ]
                            tables.append(ExtractedTable(
                                headers=headers,
                                rows=rows,
                                page_number=page_idx + 1,
                                table_index=table_idx,
                            ))

        full_text = "\n\n".join(pages_text)
        processing_time = int((time.time() - start_time) * 1000)

        metadata = DocumentMetadata(
            id=hashlib.md5(str(path).encode()).hexdigest()[:16],
            filename=path.name,
            document_type=DocumentType.PDF,
            file_size=path.stat().st_size,
            page_count=page_count,
            created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            has_text=bool(full_text.strip()),
            has_tables=bool(tables),
            processing_time_ms=processing_time,
        )

        return DocumentAnalysis(
            success=True,
            metadata=metadata,
            text=full_text,
            pages=pages_text,
            tables=tables,
        )

    async def _analyze_pdf_with_pypdf2(
        self,
        path: Path,
        max_pages: Optional[int],
        start_time: float,
    ) -> DocumentAnalysis:
        """Analyze PDF using PyPDF2."""
        import PyPDF2
        import time

        pages_text = []

        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            page_count = len(reader.pages)
            pages_to_process = reader.pages[:max_pages] if max_pages else reader.pages

            for page in pages_to_process:
                text = page.extract_text() or ""
                pages_text.append(text)

        full_text = "\n\n".join(pages_text)
        processing_time = int((time.time() - start_time) * 1000)

        metadata = DocumentMetadata(
            id=hashlib.md5(str(path).encode()).hexdigest()[:16],
            filename=path.name,
            document_type=DocumentType.PDF,
            file_size=path.stat().st_size,
            page_count=page_count,
            created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            has_text=bool(full_text.strip()),
            has_tables=False,  # PyPDF2 doesn't extract tables
            processing_time_ms=processing_time,
        )

        return DocumentAnalysis(
            success=True,
            metadata=metadata,
            text=full_text,
            pages=pages_text,
        )

    async def analyze_spreadsheet(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        detect_headers: bool = True,
    ) -> DocumentAnalysis:
        """Parse Excel/CSV with column detection.

        Args:
            file_path: Path to spreadsheet file
            sheet_name: Specific sheet to read (Excel only)
            detect_headers: Auto-detect header row

        Returns:
            DocumentAnalysis with extracted data
        """
        import time
        start_time = time.time()

        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return DocumentAnalysis(
                success=False,
                error=f"File not found: {file_path}"
            )

        ext = path.suffix.lower()

        try:
            if ext == ".csv":
                return await self._analyze_csv(path, detect_headers, start_time)
            elif ext in (".xlsx", ".xls"):
                return await self._analyze_excel(path, sheet_name, detect_headers, start_time)
            else:
                return DocumentAnalysis(
                    success=False,
                    error=f"Unsupported spreadsheet format: {ext}"
                )
        except Exception as e:
            logger.error(f"Spreadsheet analysis failed: {e}")
            return DocumentAnalysis(
                success=False,
                error=f"Spreadsheet analysis failed: {str(e)}"
            )

    async def _analyze_csv(
        self,
        path: Path,
        detect_headers: bool,
        start_time: float,
    ) -> DocumentAnalysis:
        """Analyze CSV file."""
        import time

        tables = []
        text_parts = []

        with open(path, "r", encoding="utf-8", errors="replace") as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows = list(reader)

        if rows:
            if detect_headers:
                headers = rows[0]
                data_rows = rows[1:]
            else:
                headers = [f"Column {i+1}" for i in range(len(rows[0]))]
                data_rows = rows

            tables.append(ExtractedTable(
                headers=headers,
                rows=data_rows,
                page_number=1,
                table_index=0,
            ))

            # Generate text representation
            text_parts.append(f"CSV with {len(data_rows)} rows and {len(headers)} columns")
            text_parts.append(f"Columns: {', '.join(headers)}")

        processing_time = int((time.time() - start_time) * 1000)

        metadata = DocumentMetadata(
            id=hashlib.md5(str(path).encode()).hexdigest()[:16],
            filename=path.name,
            document_type=DocumentType.CSV,
            file_size=path.stat().st_size,
            page_count=1,
            created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            has_text=True,
            has_tables=True,
            processing_time_ms=processing_time,
        )

        return DocumentAnalysis(
            success=True,
            metadata=metadata,
            text="\n".join(text_parts),
            tables=tables,
        )

    async def _analyze_excel(
        self,
        path: Path,
        sheet_name: Optional[str],
        detect_headers: bool,
        start_time: float,
    ) -> DocumentAnalysis:
        """Analyze Excel file."""
        import time

        try:
            import openpyxl
        except ImportError:
            return DocumentAnalysis(
                success=False,
                error="openpyxl not installed. Run: pip install openpyxl"
            )

        tables = []
        text_parts = []

        wb = openpyxl.load_workbook(path, data_only=True)

        sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames

        for idx, name in enumerate(sheets_to_process):
            if name not in wb.sheetnames:
                continue

            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # Filter out completely empty rows
            rows = [row for row in rows if any(cell is not None for cell in row)]

            if not rows:
                continue

            if detect_headers:
                headers = [str(cell) if cell is not None else "" for cell in rows[0]]
                data_rows = [
                    [str(cell) if cell is not None else "" for cell in row]
                    for row in rows[1:]
                ]
            else:
                headers = [f"Column {i+1}" for i in range(len(rows[0]))]
                data_rows = [
                    [str(cell) if cell is not None else "" for cell in row]
                    for row in rows
                ]

            tables.append(ExtractedTable(
                headers=headers,
                rows=data_rows,
                page_number=idx + 1,
                table_index=idx,
            ))

            text_parts.append(f"Sheet '{name}': {len(data_rows)} rows, {len(headers)} columns")
            text_parts.append(f"Columns: {', '.join(headers)}")

        wb.close()

        processing_time = int((time.time() - start_time) * 1000)

        metadata = DocumentMetadata(
            id=hashlib.md5(str(path).encode()).hexdigest()[:16],
            filename=path.name,
            document_type=DocumentType.EXCEL,
            file_size=path.stat().st_size,
            page_count=len(tables),
            created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            has_text=True,
            has_tables=True,
            processing_time_ms=processing_time,
        )

        return DocumentAnalysis(
            success=True,
            metadata=metadata,
            text="\n\n".join(text_parts),
            tables=tables,
        )

    async def analyze_image_document(
        self,
        file_path: str,
        document_type: str = "other",
        extract_fields: Optional[List[str]] = None,
    ) -> DocumentAnalysis:
        """OCR for scanned documents using Claude Vision API.

        Args:
            file_path: Path to image file
            document_type: Type hint (invoice, receipt, form, etc.)
            extract_fields: Specific fields to extract

        Returns:
            DocumentAnalysis with extracted text
        """
        import time
        start_time = time.time()

        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return DocumentAnalysis(
                success=False,
                error=f"File not found: {file_path}"
            )

        # Read image
        try:
            with open(path, "rb") as f:
                image_data = f.read()
        except Exception as e:
            return DocumentAnalysis(
                success=False,
                error=f"Failed to read image: {str(e)}"
            )

        # Validate image type
        media_type, _ = mimetypes.guess_type(str(path))
        if not media_type or not media_type.startswith("image/"):
            return DocumentAnalysis(
                success=False,
                error=f"Not a valid image file: {path.name}"
            )

        # Build prompt
        prompt = self._build_ocr_prompt(document_type, extract_fields)

        # Call Claude Vision
        try:
            image_base64 = base64.standard_b64encode(image_data).decode("utf-8")

            response = await self.client.messages.create(
                model=self.VISION_MODEL,
                max_tokens=self.MAX_TOKENS,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": image_base64,
                                },
                            },
                            {
                                "type": "text",
                                "text": prompt,
                            },
                        ],
                    }
                ],
            )

            extracted_text = ""
            for block in response.content:
                if block.type == "text":
                    extracted_text += block.text

            processing_time = int((time.time() - start_time) * 1000)

            metadata = DocumentMetadata(
                id=hashlib.md5(str(path).encode()).hexdigest()[:16],
                filename=path.name,
                document_type=DocumentType.IMAGE,
                file_size=len(image_data),
                page_count=1,
                created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
                processed_at=datetime.utcnow().isoformat(),
                has_text=bool(extracted_text.strip()),
                has_images=True,
                processing_time_ms=processing_time,
                tokens_used=response.usage.input_tokens + response.usage.output_tokens,
            )

            return DocumentAnalysis(
                success=True,
                metadata=metadata,
                text=extracted_text,
                pages=[extracted_text],
            )

        except Exception as e:
            logger.error(f"Image OCR failed: {e}")
            return DocumentAnalysis(
                success=False,
                error=f"Image OCR failed: {str(e)}"
            )

    def _build_ocr_prompt(
        self,
        document_type: str,
        extract_fields: Optional[List[str]] = None,
    ) -> str:
        """Build OCR prompt based on document type."""
        base_prompts = {
            "invoice": (
                "Extract all text from this invoice image. Include:\n"
                "1. Invoice number and date\n"
                "2. Seller/vendor information\n"
                "3. Buyer information\n"
                "4. All line items with descriptions and amounts\n"
                "5. Subtotal, taxes, and total\n"
                "6. Payment terms and due date"
            ),
            "receipt": (
                "Extract all text from this receipt. Include:\n"
                "1. Store/merchant name and location\n"
                "2. Date and time\n"
                "3. All items purchased with prices\n"
                "4. Subtotal, tax, and total\n"
                "5. Payment method"
            ),
            "form": (
                "Extract all text from this form. Include:\n"
                "1. Form title and type\n"
                "2. All field labels and their filled-in values\n"
                "3. Checkboxes and their states\n"
                "4. Signatures and dates"
            ),
            "contract": (
                "Extract all text from this contract. Include:\n"
                "1. Contract title and parties\n"
                "2. Key terms and conditions\n"
                "3. Dates and deadlines\n"
                "4. Financial terms\n"
                "5. Signatures"
            ),
            "other": (
                "Extract all visible text from this document image.\n"
                "Preserve the structure and layout as much as possible.\n"
                "Include headers, body text, tables, and any other text."
            ),
        }

        prompt = base_prompts.get(document_type, base_prompts["other"])

        if extract_fields:
            prompt += f"\n\nSpecifically extract these fields: {', '.join(extract_fields)}"

        prompt += "\n\nProvide the extracted text in a clear, structured format."

        return prompt

    async def analyze_text_file(
        self,
        file_path: str,
    ) -> DocumentAnalysis:
        """Analyze plain text or markdown files.

        Args:
            file_path: Path to text file

        Returns:
            DocumentAnalysis with content
        """
        import time
        start_time = time.time()

        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return DocumentAnalysis(
                success=False,
                error=f"File not found: {file_path}"
            )

        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
        except Exception as e:
            return DocumentAnalysis(
                success=False,
                error=f"Failed to read file: {str(e)}"
            )

        processing_time = int((time.time() - start_time) * 1000)
        doc_type = DocumentType.MARKDOWN if path.suffix.lower() in (".md", ".markdown") else DocumentType.TEXT

        metadata = DocumentMetadata(
            id=hashlib.md5(str(path).encode()).hexdigest()[:16],
            filename=path.name,
            document_type=doc_type,
            file_size=path.stat().st_size,
            page_count=1,
            created_at=datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            has_text=bool(text.strip()),
            processing_time_ms=processing_time,
        )

        return DocumentAnalysis(
            success=True,
            metadata=metadata,
            text=text,
            pages=[text],
        )

    async def analyze(
        self,
        file_path: str,
        **kwargs,
    ) -> DocumentAnalysis:
        """Analyze any supported document type.

        Auto-detects the document type and routes to appropriate analyzer.

        Args:
            file_path: Path to document
            **kwargs: Additional arguments passed to specific analyzer

        Returns:
            DocumentAnalysis with results
        """
        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            return DocumentAnalysis(
                success=False,
                error=f"File not found: {file_path}"
            )

        doc_type = self.detect_document_type(path.name)

        if doc_type == DocumentType.PDF:
            return await self.analyze_pdf(str(path), **kwargs)
        elif doc_type == DocumentType.EXCEL:
            return await self.analyze_spreadsheet(str(path), **kwargs)
        elif doc_type == DocumentType.CSV:
            return await self.analyze_spreadsheet(str(path), **kwargs)
        elif doc_type == DocumentType.IMAGE:
            return await self.analyze_image_document(str(path), **kwargs)
        elif doc_type in (DocumentType.TEXT, DocumentType.MARKDOWN):
            return await self.analyze_text_file(str(path))
        else:
            return DocumentAnalysis(
                success=False,
                error=f"Unsupported document type: {path.suffix}"
            )

    async def summarize_document(
        self,
        content: str,
        max_length: int = 500,
    ) -> Dict[str, Any]:
        """Generate an executive summary of document content.

        Args:
            content: Document text content
            max_length: Maximum summary length in characters

        Returns:
            Summary result with key points
        """
        if not content.strip():
            return {
                "success": False,
                "error": "No content to summarize"
            }

        # Truncate content if too long
        max_content = 50000
        if len(content) > max_content:
            content = content[:max_content] + "\n\n[Content truncated...]"

        prompt = self.SUMMARY_PROMPT.format(content=content)

        try:
            response = await self.client.messages.create(
                model=self.ANALYSIS_MODEL,
                max_tokens=1500,
                messages=[{"role": "user", "content": prompt}],
            )

            summary_text = response.content[0].text if response.content else ""

            # Extract key points from the response
            key_points = []
            for line in summary_text.split("\n"):
                line = line.strip()
                if line.startswith("- ") or line.startswith("* "):
                    key_points.append(line[2:])

            return {
                "success": True,
                "summary": summary_text,
                "key_points": key_points,
                "tokens_used": response.usage.input_tokens + response.usage.output_tokens,
            }

        except Exception as e:
            logger.error(f"Document summarization failed: {e}")
            return {
                "success": False,
                "error": f"Summarization failed: {str(e)}"
            }

    async def extract_data(
        self,
        content: str,
        extraction_types: List[str],
        custom_fields: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Pull specific fields from document content.

        Args:
            content: Document text content
            extraction_types: Types of data to extract (dates, amounts, names, etc.)
            custom_fields: Custom field names to extract

        Returns:
            Extracted data
        """
        if not content.strip():
            return {
                "success": False,
                "error": "No content to extract from"
            }

        # Build extraction specification
        extraction_spec = []

        for ext_type in extraction_types:
            if ext_type == "dates":
                extraction_spec.append("- All dates mentioned (in ISO format: YYYY-MM-DD)")
            elif ext_type == "amounts":
                extraction_spec.append("- All monetary amounts with currency")
            elif ext_type == "names":
                extraction_spec.append("- All person names mentioned")
            elif ext_type == "emails":
                extraction_spec.append("- All email addresses")
            elif ext_type == "phones":
                extraction_spec.append("- All phone numbers")
            elif ext_type == "addresses":
                extraction_spec.append("- All physical/mailing addresses")
            elif ext_type == "tables":
                extraction_spec.append("- Any tabular data as arrays")

        if custom_fields:
            for field in custom_fields:
                extraction_spec.append(f"- {field}")

        extraction_text = "\n".join(extraction_spec)

        # Truncate content if too long
        max_content = 50000
        if len(content) > max_content:
            content = content[:max_content] + "\n\n[Content truncated...]"

        prompt = self.EXTRACTION_PROMPT.format(
            extraction_type=extraction_text,
            content=content,
        )

        try:
            response = await self.client.messages.create(
                model=self.ANALYSIS_MODEL,
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}],
            )

            response_text = response.content[0].text if response.content else "{}"

            # Parse JSON from response
            try:
                # Find JSON in response
                start = response_text.find("{")
                end = response_text.rfind("}") + 1
                if start >= 0 and end > start:
                    extracted = json.loads(response_text[start:end])
                else:
                    extracted = {"raw_response": response_text}
            except json.JSONDecodeError:
                extracted = {"raw_response": response_text}

            return {
                "success": True,
                "extracted_data": extracted,
                "tokens_used": response.usage.input_tokens + response.usage.output_tokens,
            }

        except Exception as e:
            logger.error(f"Data extraction failed: {e}")
            return {
                "success": False,
                "error": f"Extraction failed: {str(e)}"
            }

    async def search_in_document(
        self,
        content: str,
        query: str,
    ) -> Dict[str, Any]:
        """Search for specific text or concepts in document.

        Args:
            content: Document text content
            query: Search query

        Returns:
            Search results with matches and context
        """
        if not content.strip():
            return {
                "success": False,
                "error": "No content to search"
            }

        # First, do simple text search
        simple_matches = []
        query_lower = query.lower()
        lines = content.split("\n")

        for i, line in enumerate(lines):
            if query_lower in line.lower():
                # Get context (2 lines before and after)
                start = max(0, i - 2)
                end = min(len(lines), i + 3)
                context = "\n".join(lines[start:end])
                simple_matches.append({
                    "line_number": i + 1,
                    "match": line.strip(),
                    "context": context,
                })

        # Use AI for semantic search
        max_content = 50000
        if len(content) > max_content:
            content = content[:max_content] + "\n\n[Content truncated...]"

        prompt = self.SEARCH_PROMPT.format(query=query, content=content)

        try:
            response = await self.client.messages.create(
                model=self.ANALYSIS_MODEL,
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}],
            )

            ai_results = response.content[0].text if response.content else ""

            return {
                "success": True,
                "query": query,
                "exact_matches": simple_matches[:20],  # Limit exact matches
                "semantic_results": ai_results,
                "match_count": len(simple_matches),
                "tokens_used": response.usage.input_tokens + response.usage.output_tokens,
            }

        except Exception as e:
            # Return just simple matches if AI fails
            return {
                "success": True,
                "query": query,
                "exact_matches": simple_matches[:20],
                "semantic_results": None,
                "match_count": len(simple_matches),
                "error": f"AI search failed: {str(e)}",
            }

    async def compare_documents(
        self,
        content1: str,
        content2: str,
        name1: str = "Version 1",
        name2: str = "Version 2",
    ) -> Dict[str, Any]:
        """Diff two document versions.

        Args:
            content1: First document content
            content2: Second document content
            name1: Label for first document
            name2: Label for second document

        Returns:
            Comparison results with changes identified
        """
        if not content1.strip() and not content2.strip():
            return {
                "success": False,
                "error": "Both documents are empty"
            }

        # Simple diff stats
        lines1 = set(content1.strip().split("\n"))
        lines2 = set(content2.strip().split("\n"))

        added = lines2 - lines1
        removed = lines1 - lines2
        unchanged = lines1 & lines2

        # Truncate for AI comparison
        max_content = 25000
        c1 = content1[:max_content] if len(content1) > max_content else content1
        c2 = content2[:max_content] if len(content2) > max_content else content2

        if len(content1) > max_content or len(content2) > max_content:
            c1 += "\n\n[Content truncated...]"
            c2 += "\n\n[Content truncated...]"

        prompt = self.COMPARE_PROMPT.format(content1=c1, content2=c2)

        try:
            response = await self.client.messages.create(
                model=self.ANALYSIS_MODEL,
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}],
            )

            comparison = response.content[0].text if response.content else ""

            return {
                "success": True,
                "version1": name1,
                "version2": name2,
                "stats": {
                    "lines_added": len(added),
                    "lines_removed": len(removed),
                    "lines_unchanged": len(unchanged),
                },
                "comparison": comparison,
                "tokens_used": response.usage.input_tokens + response.usage.output_tokens,
            }

        except Exception as e:
            # Return basic stats if AI fails
            return {
                "success": True,
                "version1": name1,
                "version2": name2,
                "stats": {
                    "lines_added": len(added),
                    "lines_removed": len(removed),
                    "lines_unchanged": len(unchanged),
                },
                "comparison": None,
                "error": f"AI comparison failed: {str(e)}",
            }


# ============ AI TOOL DEFINITIONS ============

DOCUMENT_TOOLS = [
    {
        "name": "analyze_document_file",
        "description": "Analyze an uploaded document file. Supports PDF, Excel, CSV, images (scanned documents), and text files. Extracts text, tables, and structure.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document file"
                },
                "document_type": {
                    "type": "string",
                    "enum": ["pdf", "excel", "csv", "image", "text", "auto"],
                    "description": "Type of document. Use 'auto' for automatic detection."
                },
                "extract_tables": {
                    "type": "boolean",
                    "description": "Whether to extract tables from the document (default: true)"
                }
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "extract_from_document",
        "description": "Extract specific data from a document. Can pull dates, amounts, names, emails, phone numbers, addresses, or custom fields. Example: 'get the total from this invoice'",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document file"
                },
                "extraction_types": {
                    "type": "array",
                    "items": {
                        "type": "string",
                        "enum": ["dates", "amounts", "names", "emails", "phones", "addresses", "tables"]
                    },
                    "description": "Types of data to extract"
                },
                "custom_fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Custom field names to extract (e.g., 'invoice number', 'due date', 'total')"
                }
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "summarize_document_file",
        "description": "Generate an executive summary of a document. Provides key points, document type identification, and important extracted information.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document file"
                },
                "max_length": {
                    "type": "integer",
                    "description": "Maximum summary length in characters (default: 500)"
                }
            },
            "required": ["file_path"]
        }
    },
    {
        "name": "compare_document_versions",
        "description": "Compare two versions of a document and identify changes. Shows added, removed, and modified content.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path_1": {
                    "type": "string",
                    "description": "Path to the first (older) document version"
                },
                "file_path_2": {
                    "type": "string",
                    "description": "Path to the second (newer) document version"
                },
                "name_1": {
                    "type": "string",
                    "description": "Label for first version (default: 'Version 1')"
                },
                "name_2": {
                    "type": "string",
                    "description": "Label for second version (default: 'Version 2')"
                }
            },
            "required": ["file_path_1", "file_path_2"]
        }
    },
    {
        "name": "search_in_document_file",
        "description": "Search for specific text or concepts in a document. Finds exact matches and semantically related content.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Path to the document file"
                },
                "query": {
                    "type": "string",
                    "description": "Search query - what to look for in the document"
                }
            },
            "required": ["file_path", "query"]
        }
    },
]


# ============ MODULE-LEVEL CONVENIENCE FUNCTIONS ============

# Singleton instances
_document_store: Optional[DocumentStore] = None
_document_analyzer: Optional[DocumentAnalyzer] = None


def get_document_store() -> DocumentStore:
    """Get or create singleton document store."""
    global _document_store
    if _document_store is None:
        _document_store = DocumentStore()
    return _document_store


def get_document_analyzer() -> DocumentAnalyzer:
    """Get or create singleton document analyzer."""
    global _document_analyzer
    if _document_analyzer is None:
        _document_analyzer = DocumentAnalyzer()
    return _document_analyzer


async def analyze_document(
    file_path: str,
    **kwargs,
) -> DocumentAnalysis:
    """Analyze any supported document.

    Convenience function that uses the singleton analyzer.
    """
    analyzer = get_document_analyzer()
    return await analyzer.analyze(file_path, **kwargs)


async def summarize_document(
    file_path: str,
    max_length: int = 500,
) -> Dict[str, Any]:
    """Summarize a document file.

    Convenience function that analyzes then summarizes.
    """
    analyzer = get_document_analyzer()

    # First analyze the document
    analysis = await analyzer.analyze(file_path)
    if not analysis.success:
        return {
            "success": False,
            "error": analysis.error
        }

    # Then summarize
    return await analyzer.summarize_document(analysis.text, max_length)


async def extract_from_document(
    file_path: str,
    extraction_types: List[str],
    custom_fields: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Extract specific data from a document.

    Convenience function that analyzes then extracts.
    """
    analyzer = get_document_analyzer()

    # First analyze the document
    analysis = await analyzer.analyze(file_path)
    if not analysis.success:
        return {
            "success": False,
            "error": analysis.error
        }

    # Then extract
    return await analyzer.extract_data(analysis.text, extraction_types, custom_fields)


async def search_in_document(
    file_path: str,
    query: str,
) -> Dict[str, Any]:
    """Search in a document.

    Convenience function that analyzes then searches.
    """
    analyzer = get_document_analyzer()

    # First analyze the document
    analysis = await analyzer.analyze(file_path)
    if not analysis.success:
        return {
            "success": False,
            "error": analysis.error
        }

    # Then search
    return await analyzer.search_in_document(analysis.text, query)


async def compare_documents(
    file_path_1: str,
    file_path_2: str,
    name_1: str = "Version 1",
    name_2: str = "Version 2",
) -> Dict[str, Any]:
    """Compare two document versions.

    Convenience function that analyzes both documents then compares.
    """
    analyzer = get_document_analyzer()

    # Analyze both documents
    analysis1 = await analyzer.analyze(file_path_1)
    if not analysis1.success:
        return {
            "success": False,
            "error": f"Error analyzing {name_1}: {analysis1.error}"
        }

    analysis2 = await analyzer.analyze(file_path_2)
    if not analysis2.success:
        return {
            "success": False,
            "error": f"Error analyzing {name_2}: {analysis2.error}"
        }

    # Compare
    return await analyzer.compare_documents(
        analysis1.text,
        analysis2.text,
        name_1,
        name_2,
    )
